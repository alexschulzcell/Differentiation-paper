# -*- coding: utf-8 -*-
# =============================================================================
# EXPLORATORY analysis "aggregated direction replication" (paper-impact module)
# =============================================================================
# Question: while no single independent study reaches the pre-fixed MDE80
# threshold of route A, a consistent directional tendency underlies them.
# The claim tested in pooled form: across all biological units of the six
# independent S12 route-A datasets, do the modular genes reproduce the S5
# direction 'ri' IN THE MAJORITY of observations?
#
# Status: EXPLORATORY (not in PRAEREG_S12 §2-6; there, route A remains
# reported individually and unchanged). Fixed gene list (S5, 173 genes), no
# new thresholds, no post-hoc selection; GSE37521 stays included. GSE200492
# (S11 anchor) is NOT pooled (PRAEREG_S12 rule "GSE200492 must not be counted
# again as an independent success").
#
# Statistic: observation = (dataset, gene): sign( mean of unit diffs )
# against ri. Pool over datasets; denominator = sum of measurable genes.
# Null: 10,000 rounds in which, per dataset, the complete unit columns are
# sign-flipped jointly (structure preserved). MDE80 analogous to the project
# standard = null + 2.80*null SD. Leave-one-dataset-out.
# Seed: 20260824 (exploratory).
# =============================================================================
from __future__ import annotations
import gzip
import io
import re
import subprocess
from pathlib import Path
import numpy as np
import pandas as pd

# --- path parameters (2026-08-23) ------------------------------------------
# Previously a hard-coded path pointed to ".../DFG Antrag/
# Scherenpaper_Folgeprojekt" -- a Windows junction, not clonable. Two
# different trees sat in this one variable:
#   * the session folders 25_Orthogonal_S11 / 26_Orthogonal_S12 -- they live
#     under `Paper v2/_archiv/Sitzungen/`. The audit of 2026-08-22 had listed
#     them as "missing"; they were just stored elsewhere.
#   * `derived_data/manuscript`, `derived_data/reference_tables` -- these live under `Paper v2` itself.
# Both are now separate and overridable via environment variables.
import os, pathlib
_env = os.environ.get("PAPER_V2_ROOT")
WURZEL = (pathlib.Path(_env) if _env
          else pathlib.Path(__file__).resolve().parents[1])
SITZUNGEN = pathlib.Path(os.environ.get(
    "SCHERENPAPER_SITZUNGEN", str(WURZEL / "_archiv" / "Sitzungen")))
DATA = SITZUNGEN / "26_Orthogonal_S12" / "data_raw"
OUT = WURZEL / "derived_data" / "manuscript"
SEED = 20260824
N_PERM = 10_000
TAU = WURZEL / "derived_data" / "reference_tables" / "S5_konvergente_gene.csv"


def module_table() -> pd.DataFrame:
    tau = pd.read_csv(TAU, dtype={"gen": str, "ri": int, "symbol": str})
    assert len(tau) == 173 and set(tau["ri"]) == {-1, 1}
    return tau


# --------------------------------------------------------------------------
# Loader: data loading 1:1 after s12_route_a.py (loading only; no kern()).
# --------------------------------------------------------------------------
def diff_gse185951():
    raw = pd.read_csv(DATA / "GSE185951" / "GSE185951_hMSCdonor_normalized_counts.txt.gz",
                      sep="\t", compression="gzip", index_col=0)
    raw.index = raw.index.astype(str).str.replace(r"\..*$", "", regex=True)
    raw = raw.groupby(level=0, sort=False).mean()
    pat = re.compile(r"^hMSC_(?P<donor>\d+)_day(?P<day>0|7|14)$")
    keep = [(c, int(m.group("donor"))) for c in raw.columns if (m := pat.match(c))
            and m.group("day") in ("0", "7")]
    donors = sorted({d for _, d in keep})
    cols = {f"hMSC_{d}_day{x}" for d in donors for x in ("0", "7")}
    mat = raw[sorted(cols)]
    return pd.DataFrame({f"donor_{d}": mat[f"hMSC_{d}_day7"] - mat[f"hMSC_{d}_day0"]
                         for d in donors}), "ensg"


def diff_gse161176():
    raw = pd.read_csv(DATA / "GSE161176" / "GSE161176_genes_counts.txt.gz",
                      sep="\t", compression="gzip", index_col=1)
    raw = raw.apply(pd.to_numeric, errors="coerce")
    raw = raw.drop(columns=[c for c in raw.columns if str(c) in ("", "nan") or str(c).startswith("Unnamed")])
    raw.columns = [str(c) for c in raw.columns]
    soft = DATA / "_kandidaten_metadaten" / "GSE161176_family.soft.gz"
    with gzip.open(soft, "rt", encoding="utf-8", errors="replace") as fh:
        titles = [m.group(1) for m in re.finditer(r"^!Sample_title = (.+)$", fh.read(), re.M)]
    pat = re.compile(
        r"^(?P<kind>BMSC|ACh) Donor (?P<d>\d+)(?: Day 0)?"
        r"(?:, (?P<days>\d+)-days? TGF-B\d+ exposure, analyzed on Day (?P<day>\d+))?"
        r"_rep ?(?P<r>\d+)$")
    meta = []
    for col, t in zip(raw.columns, titles):
        m = pat.match(t)
        if not m:
            raise ValueError(t)
        g = m.groupdict()
        meta.append({"col": col, "kind": g["kind"], "donor": g["d"],
                     "day": g["day"] or "0"})
    meta = pd.DataFrame(meta)
    b = meta[meta.kind == "BMSC"]
    raw = raw.loc[:, b.col]
    rows = []
    for donor in sorted(b.donor.unique()):
        d0 = raw[b[(b.donor == donor) & (b.day == "0")].col.values].mean(axis=1)
        d1 = raw[b[(b.donor == donor) & (b.day == "1")].col.values].mean(axis=1)
        rows.append(pd.Series(d1 - d0, name=f"BMSC_d{donor}"))
    return pd.concat(rows, axis=1), "symbol"


def diff_gse113253():
    df = pd.read_csv(DATA / "GSE113253" / "GSE113253_RNA_HumanPrimaryStromalCells.txt.gz",
                     sep="\t", compression="gzip", index_col=0)
    ann = df["Annotation.Divergence"].str.split("|").str[0]
    keep = [c for c in df.columns if re.match(r"^(WAT|MUS|BM)_(Msc|Ob)_(1|2|3)$", c)]
    d = df[keep].copy()
    d["symbol"] = ann
    d = d.groupby("symbol").mean(numeric_only=True)
    m2 = pd.DataFrame({f"{c.split('_')[0]}_{c.split('_')[2]}__{c.split('_')[1]}": d[c]
                       for c in d.columns})
    cols = sorted({c.split("__")[0] for c in m2.columns})
    for unit in cols:
        m2[f"{unit}__diff"] = m2[f"{unit}__Ob"] - m2[f"{unit}__Msc"]
    diff = m2[[f"{u}__diff" for u in cols]]
    diff.columns = cols
    return diff, "symbol"


def diff_gse210984():
    cols = pd.read_csv(DATA / "GSE210984" / "GSE210984_gene_fpkm.txt.gz",
                       sep="\t", compression="gzip", nrows=0).columns
    samp = [c for c in cols if re.match(r"^(MSC|C7)_\d+$", str(c))]
    raw = pd.read_csv(DATA / "GSE210984" / "GSE210984_gene_fpkm.txt.gz",
                      sep="\t", compression="gzip", usecols=["ID"] + samp)
    raw = raw.set_index("ID")
    raw.index = raw.index.astype(str).str.replace(r"\..*$", "", regex=True)
    raw = raw.groupby(level=0, sort=False).mean()
    raw = raw[samp]
    m = pd.DataFrame({f"rep{j}__{arm}": raw[c]
                      for arm in ["MSC", "C7"]
                      for j, c in enumerate([x for x in samp if x.startswith(arm + "_")], start=1)})
    return pd.DataFrame({f"rep{i}": m[f"rep{i}__C7"] - m[f"rep{i}__MSC"]
                         for i in range(1, 4)}), "ensg"


def diff_gse202080():
    import openpyxl
    tar = DATA / "GSE202080" / "GSE202080_RAW.tar"
    entries = subprocess.run(["tar", "-tf", str(tar)], capture_output=True, text=True).stdout.splitlines()
    pat = re.compile(r"^GSM6094\d{3}_(?P<src>[A-Za-z0-9_]+?)_(?P<stage>[1-4])_OBs_differentiation\.xls\.gz$")
    stage_n = {"1": "MSCs", "2": "Pre-OBs", "3": "Mat-OBs", "4": "OBs"}
    frames = {}
    for entry in entries:
        m = pat.match(entry)
        if not m:
            continue
        out = subprocess.run(["tar", "-xOf", str(tar), entry], capture_output=True)
        wb = openpyxl.load_workbook(io.BytesIO(gzip.decompress(out.stdout)), read_only=True)
        ws = wb[wb.sheetnames[0]]
        it = ws.iter_rows(values_only=True)
        next(it)
        df = pd.DataFrame(it, columns=["gene_id", "alias", "fpkm"])
        df["fpkm"] = pd.to_numeric(df["fpkm"], errors="coerce")
        frames[(m.group("src"), stage_n[m.group("stage")])] = df.set_index("alias")["fpkm"]
        wb.close()
    srcs = [s for s in sorted({s for (s, t) in frames})
            if (s, "Pre-OBs") in frames and (s, "MSCs") in frames]
    diff = pd.DataFrame({s: frames[(s, "Pre-OBs")] - frames[(s, "MSCs")] for s in srcs})
    return diff.groupby(level=0, sort=False).mean(), "symbol"


def diff_gse37521():
    p = DATA / "GSE37521" / "GSE37521_L127_RPM.tab.txt.gz"
    with gzip.open(p, "rt", encoding="utf-8", errors="replace") as fh:
        lines = fh.readlines()
    meta = {}
    for line in lines:
        cells = line.rstrip("\n").split("\t")
        for j, c in enumerate(cells):
            tag = c.strip().rstrip(":") if c.strip().endswith(":") else ""
            if tag in ("Barcode", "Sample", "Source", "Destination", "Day", "Patient"):
                meta[tag] = [x.strip() for x in cells[j + 1:] if x.strip() != ""]
    n = len(meta["Patient"])
    use = list(range(n))
    rows = []
    for line in lines:
        cells = line.rstrip("\n").split("\t")
        if len(cells) < 10 + 96:
            continue
        gene = cells[0].strip()
        if not gene or gene in ("SingleRead", "Feature"):
            continue
        rows.append([gene] + [cells[10 + i] for i in use])
    m = pd.DataFrame(rows)
    m.columns = ["Feature"] + [f"well_{i}" for i in use]
    for c in m.columns[1:]:
        m[c] = pd.to_numeric(m[c], errors="coerce")
    m = m.groupby("Feature").mean(numeric_only=True)
    ni = [i for i in use if meta["Destination"][i] == "NI" and meta["Day"][i] == "d0"]
    diff = pd.DataFrame()
    for pat in sorted(set(meta["Patient"][i] for i in use)):
        oday = sorted({meta["Day"][i] for i in use
                       if meta["Patient"][i] == pat and meta["Destination"][i] == "O"},
                      key=lambda d: int(d[1:]))
        if not oday:
            continue
        ost = [i for i in use
               if meta["Patient"][i] == pat and meta["Destination"][i] == "O" and meta["Day"][i] == oday[0]]
        if not ni or not ost:
            continue
        diff[f"{pat}_O{oday[0]}"] = m.iloc[:, ost].mean(axis=1) - m.iloc[:, ni].mean(axis=1)
    return diff.groupby("Feature").mean(numeric_only=True), "symbol"


# --------------------------------------------------------------------------
# Joint evaluation in main(): observations per dataset mirror the statistic
# from s12_route_a.metric (sign(unit-mean diff) vs ri, complete cases);
# assertions against s12_route_a_summary.csv prevent any deviation.
# --------------------------------------------------------------------------
def main() -> int:
    OUT.mkdir(parents=True, exist_ok=True)
    tau = module_table()
    rng = np.random.default_rng(SEED)

    specs = [
        ("GSE185951", diff_gse185951, "ensg"),
        ("GSE161176", diff_gse161176, "symbol"),
        ("GSE113253", diff_gse113253, "symbol"),
        ("GSE210984", diff_gse210984, "ensg"),
        ("GSE202080", diff_gse202080, "symbol"),
        ("GSE37521", diff_gse37521, "symbol"),
    ]
    ref_csv = SITZUNGEN / "26_Orthogonal_S12" / "derived_data" / "s12_route_a_summary.csv"

    mats, genes = {}, {}
    for name, fn, key_type in specs:
        print(f"loading {name} ...", flush=True)
        diff, _ = fn()
        mp = tau.set_index("gen")["ri"] if key_type == "ensg" else tau.set_index("symbol")["ri"]
        idx = [g for g in diff.index if g in mp.index]
        d = diff.loc[idx].dropna(axis=0, how="any")
        mats[name] = d.to_numpy(dtype=float)
        genes[name] = mp.reindex(d.index).to_numpy(dtype=int)
        exp = pd.read_csv(ref_csv)
        r = exp[exp.dataset == name].iloc[0]
        assert int((np.sign(mats[name].mean(axis=1)) == genes[name]).sum()) == int(r["n_matching"])
        assert d.shape[1] == int(r["n_biological_units"])
        print(f"  {name}: units={d.shape[1]} measurable={d.shape[0]}", flush=True)

    # ---------------- observed pool ----------------
    obs_matches = sum(int((np.sign(mats[n].mean(axis=1)) == genes[n]).sum()) for n in mats)
    obs_denom = sum(mats[n].shape[0] for n in mats)
    obs_share = obs_matches / obs_denom

    # ---------------- Null: column flips per data set simultaneously ----------------
    def pooled_share() -> float:
        tot = 0
        for n in mats:
            flips = rng.choice(np.array([-1.0, 1.0]), size=mats[n].shape[1])
            tot += int((np.sign((mats[n] @ flips) / flips.size) == genes[n]).sum())
        return tot / obs_denom

    nulls = np.fromiter((pooled_share() for _ in range(N_PERM)), dtype=float, count=N_PERM)
    null_mean = float(nulls.mean())
    null_sd = float(nulls.std(ddof=1))
    z = (obs_share - null_mean) / null_sd if null_sd else float("nan")
    p_two = float(np.mean(np.abs(nulls - null_mean) >= abs(obs_share - null_mean)))
    mde80 = null_mean + 2.80 * null_sd
    p_hat = obs_matches / obs_denom
    zc = 1.959963984540054
    wil_lo = max(0.0, (p_hat + zc * zc / (2 * obs_denom) -
                       zc * np.sqrt(p_hat * (1 - p_hat) / obs_denom + zc * zc / (4 * obs_denom * obs_denom)))
                  / (1 + zc * zc / obs_denom))
    wil_hi = min(1.0, (p_hat + zc * zc / (2 * obs_denom) +
                       zc * np.sqrt(p_hat * (1 - p_hat) / obs_denom + zc * zc / (4 * obs_denom * obs_denom)))
                  / (1 + zc * zc / obs_denom))

    # ---------------- Leave-one-dataset-out ----------------
    loo_rows = []
    for drop in mats:
        dom = mats if False else [n for n in mats if n != drop]
        obs_m = sum(int((np.sign(mats[n].mean(axis=1)) == genes[n]).sum()) for n in dom)
        obs_d = sum(mats[n].shape[0] for n in dom)
        sub_share = obs_m / obs_d
        loo_rows.append({"omitted_dataset": drop, "pooled_share": float(sub_share),
                         "n_observations": obs_d, "n_matching": obs_m})
    loo = pd.DataFrame(loo_rows)

    summary = dict(
        analysis="explorativ_aggregierte_richtungsreplikation",
        n_datasets=len(mats), n_observations=obs_denom, n_matching=obs_matches,
        pooled_share=float(obs_share), perm_rounds=N_PERM,
        null_mean=null_mean, null_sd=null_sd, z=float(z), p_perm_two_sided=p_two,
        wilson_lo=float(wil_lo), wilson_hi=float(wil_hi),
        mde80=float(mde80), above_mde80=bool(obs_share > mde80),
        loo_min=float(loo.pooled_share.min()), loo_max=float(loo.pooled_share.max()),
        seed=SEED,
    )
    print(" | ".join(f"{k}={v}" for k, v in summary.items()))
    pd.DataFrame([summary]).to_csv(OUT / "f6_pooled_summary.csv", index=False)
    pd.DataFrame({"pooled_share_null": nulls}).to_csv(OUT / "f6_pooled_null.csv", index=False)
    loo.to_csv(OUT / "f6_pooled_loo.csv", index=False)

    forest = []
    for n, _fn, _kt in specs:
        r = pd.read_csv(ref_csv)
        r = r[r.dataset == n].iloc[0]
        forest.append(dict(dataset=n, n_units=int(r["n_biological_units"]),
                           n_measurable=int(r["n_measurable"]), share=r["share_match"],
                           wilson_lo=r["wilson_lo"], wilson_hi=r["wilson_hi"],
                           z=r["z"], p_perm=r["p_perm_two_sided"], mde80=r["mde80"],
                           above_mde80=bool(r["above_mde80"]),
                           loo_share_no_drop=float(loo[loo.omitted_dataset == n].pooled_share.iloc[0])))
    # S11 anchor: referenced only, NOT pooled.
    r11 = pd.read_csv(SITZUNGEN / "25_Orthogonal_S11" / "derived_data" / "s11_route_a_GSE200492.csv").iloc[0]
    forest.append(dict(dataset="GSE200492 (S11-Anker)", n_units=int(r11["n_biological_units"]),
                       n_measurable=int(r11["n_measurable"]), share=r11["share_match"],
                       wilson_lo=r11["wilson_lo"], wilson_hi=r11["wilson_hi"],
                       z=r11["z"], p_perm=r11["p_perm_two_sided"], mde80=r11["mde80"],
                       above_mde80=bool(r11["above_mde80"]), loo_share_no_drop=float("nan")))
    pd.DataFrame(forest).to_csv(OUT / "f6_forest.csv", index=False)
    (WURZEL / "manuscript" / "reference_implementations" / "_explorativ_aggregat_log.txt").write_text(
        "\n".join([f"{k}={v}" for k, v in summary.items()]) + "\n", encoding="utf-8")
    print("done")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
